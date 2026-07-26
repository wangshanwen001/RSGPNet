import os
import os.path as osp
import argparse
import time
import openpyxl
import torch

from mmengine.runner import Runner
from mmengine.config import Config, DictAction

import RSGPNet
import SegEarthOV3
import custom_datasets
import boundary_f1_metric


def parse_args():
    parser = argparse.ArgumentParser(
        description='CorrCLIP evaluation with MMSeg')

    parser.add_argument('config', default='./configs/cfg_potsdam_bf1.py')

    parser.add_argument('--show', action='store_true')
    parser.add_argument('--show_dir', default='./show_dir/')
    parser.add_argument('--out', type=str)

    parser.add_argument(
        '--cfg-options',
        nargs='+',
        action=DictAction)

    parser.add_argument(
        '--launcher',
        choices=['none', 'pytorch', 'slurm', 'mpi'],
        default='none')

    parser.add_argument('--local_rank', '--local-rank', type=int, default=0)

    parser.add_argument(
        '--calc-memory',
        action='store_true',
        default=True,
        help='calculate model params and GPU memory'
    )

    parser.add_argument(
        '--model-type',
        choices=['RSGPNet', 'SegEarthOV3'],
        default='RSGPNet',
        help='Model type to use for evaluation'
    )

    args = parser.parse_args()

    if 'LOCAL_RANK' not in os.environ:
        os.environ['LOCAL_RANK'] = str(args.local_rank)

    return args


def add_mfscore_to_evaluator(cfg):
    """Add mFscore to IoUMetric and add BoundaryF1Metric for mBF1."""
    evaluator = cfg.test_evaluator

    def _add_mfscore_to_iou_metric(metric_cfg):
        if not isinstance(metric_cfg, dict):
            return metric_cfg
        if metric_cfg.get('type') != 'IoUMetric':
            return metric_cfg

        if 'iou_metrics' in metric_cfg:
            if 'mFscore' not in metric_cfg['iou_metrics']:
                metric_cfg['iou_metrics'].append('mFscore')
        elif 'metrics' in metric_cfg:
            if 'mFscore' not in metric_cfg['metrics']:
                metric_cfg['metrics'].append('mFscore')
        else:
            metric_cfg['iou_metrics'] = ['mIoU', 'mFscore']
        return metric_cfg

    def _has_boundary_metric(evaluator_cfg):
        if isinstance(evaluator_cfg, dict):
            return evaluator_cfg.get('type') == 'BoundaryF1Metric'
        if isinstance(evaluator_cfg, (list, tuple)):
            return any(isinstance(x, dict) and x.get('type') == 'BoundaryF1Metric'
                       for x in evaluator_cfg)
        return False

    boundary_metric = dict(
        type='BoundaryF1Metric',
        num_classes=7,
        ignore_index=255,
        epsilon=0.02
    )

    if isinstance(evaluator, dict):
        evaluator = _add_mfscore_to_iou_metric(evaluator)
        if not _has_boundary_metric(evaluator):
            cfg.test_evaluator = [evaluator, boundary_metric]
    elif isinstance(evaluator, (list, tuple)):
        evaluator = [_add_mfscore_to_iou_metric(x) for x in evaluator]
        if not _has_boundary_metric(evaluator):
            evaluator.append(boundary_metric)
        cfg.test_evaluator = evaluator

    return cfg


def append_experiment_result(file_path, experiment_data):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if sheet['A1'].value is None:
        sheet['A1'] = 'Model'
        sheet['B1'] = 'Dataset'
        sheet['C1'] = 'aAcc'
        sheet['D1'] = 'mIoU'
        sheet['E1'] = 'mAcc'
        sheet['F1'] = 'mF1'
        sheet['G1'] = 'FPS'
        sheet['H1'] = 'mBF1'
        sheet['I1'] = 'Params(M)'
        sheet['J1'] = 'Memory(GB)'
        sheet['K1'] = 'PeakMem(GB)'

    last_row = sheet.max_row

    for index, result in enumerate(experiment_data, start=1):
        row = last_row + index

        sheet.cell(row=row, column=1, value=result.get('Model'))
        sheet.cell(row=row, column=2, value=result.get('Dataset'))
        sheet.cell(row=row, column=3, value=result.get('aAcc'))
        sheet.cell(row=row, column=4, value=result.get('mIoU'))
        sheet.cell(row=row, column=5, value=result.get('mAcc'))
        sheet.cell(row=row, column=6,
                   value=result.get('mFscore', result.get('mF1')))
        sheet.cell(row=row, column=7, value=result.get('FPS'))
        sheet.cell(row=row, column=8, value=result.get('mBF1', result.get('BF1')))
        sheet.cell(row=row, column=9, value=result.get('params(M)'))
        sheet.cell(row=row, column=10, value=result.get('memory_allocated(GB)'))
        sheet.cell(row=row, column=11, value=result.get('peak_memory(GB)'))

    workbook.save(file_path)


def get_dataset_size(runner):
    try:
        return len(runner.test_dataloader.dataset)
    except Exception:
        return None


def calculate_gpu_memory(runner):
    if not torch.cuda.is_available():
        return None

    model = runner.model
    model.eval()

    torch.cuda.empty_cache()
    torch.cuda.synchronize()

    baseline_allocated = torch.cuda.memory_allocated()
    baseline_reserved = torch.cuda.memory_reserved()

    dataloader = runner.test_dataloader

    peak_allocated = 0
    peak_reserved = 0

    with torch.no_grad():
        for data in dataloader:
            data = runner.model.data_preprocessor(data, False)

            inputs = data['inputs']
            data_samples = data.get('data_samples', None)

            if isinstance(inputs, torch.Tensor):
                inputs = inputs.cuda()

            if data_samples is not None:
                for sample in data_samples:
                    sample.to('cuda')

            torch.cuda.synchronize()

            model.predict(inputs, data_samples)

            torch.cuda.synchronize()

            current_allocated = torch.cuda.memory_allocated()
            current_reserved = torch.cuda.memory_reserved()

            peak_allocated = max(peak_allocated, current_allocated)
            peak_reserved = max(peak_reserved, current_reserved)

            break

    model_memory_allocated = peak_allocated - baseline_allocated
    model_memory_reserved = peak_reserved - baseline_reserved

    model_memory_allocated_gb = model_memory_allocated / 1e9
    model_memory_reserved_gb = model_memory_reserved / 1e9

    total_params = sum(p.numel() for p in model.parameters())
    
    if total_params == 0:
        if hasattr(model, 'processor') and hasattr(model.processor, 'model'):
            sam_model = model.processor.model
            total_params = sum(p.numel() for p in sam_model.parameters())
        elif hasattr(model, 'sam3_model'):
            total_params = sum(p.numel() for p in model.sam3_model.parameters())
    
    total_params_m = total_params / 1e6

    result = {
        'params(M)': round(total_params_m, 2),
        'memory_allocated(GB)': round(model_memory_allocated_gb, 4),
        'memory_reserved(GB)': round(model_memory_reserved_gb, 4),
        'peak_memory(GB)': round(peak_reserved / 1e9, 4)
    }

    return result


def main():
    args = parse_args()
    print(os.getcwd())

    cfg = Config.fromfile(args.config)
    cfg.launcher = args.launcher

    if args.out is not None:
        cfg.test_evaluator['output_dir'] = args.out
        cfg.test_evaluator['keep_results'] = True

    if args.cfg_options is not None:
        cfg.merge_from_dict(args.cfg_options)

    cfg.work_dir = osp.join(
        './work_dirs',
        osp.splitext(osp.basename(args.config))[0]
    )

    cfg = add_mfscore_to_evaluator(cfg)

    if args.model_type == 'RSGPNet':
        cfg.model.type = 'RSGPNetSegmentation'
        cfg.model.model_type = 'RSGPNet'
    elif args.model_type == 'SegEarthOV3':
        cfg.model.type = 'SegEarthOV3Segmentation'
        cfg.model.model_type = 'SegEarthOV3'

    runner = Runner.from_cfg(cfg)

    memory_result = None
    if runner.rank == 0 and args.calc_memory:
        print('Calculating model params and GPU memory...')
        memory_result = calculate_gpu_memory(runner)
        if memory_result is not None:
            print(f"Params(M): {memory_result['params(M)']}")
            print(f"Memory Allocated(GB): {memory_result['memory_allocated(GB)']}")
            print(f"Memory Reserved(GB): {memory_result['memory_reserved(GB)']}")
            print(f"Peak Memory(GB): {memory_result['peak_memory(GB)']}")

    if torch.cuda.is_available():
        torch.cuda.synchronize()

    start_time = time.time()

    results = runner.test()

    if torch.cuda.is_available():
        torch.cuda.synchronize()

    end_time = time.time()

    elapsed_time = end_time - start_time
    dataset_size = get_dataset_size(runner)

    fps = None
    if dataset_size is not None and elapsed_time > 0:
        fps = round(dataset_size / elapsed_time, 4)

    results.update({
        'Model': cfg.model.model_type,
        'Dataset': cfg.dataset_type,
        'FPS': fps
    })

    if memory_result is not None:
        results.update(memory_result)

    if runner.rank == 0:
        append_experiment_result('results.xlsx', [results])

    if runner.rank == 0:
        with open(os.path.join(cfg.work_dir, 'results.txt'), 'a') as f:
            f.write(os.path.basename(args.config).split('.')[0] + '\n')
            for k, v in results.items():
                f.write(k + ': ' + str(v) + '\n')

            f.write('Total test time: ' + str(round(elapsed_time, 4)) + ' s\n')

    if runner.rank == 0:
        print('Final Results:')
        for k, v in results.items():
            print(f'{k}: {v}')
        print(f'Total test time: {round(elapsed_time, 4)} s')


if __name__ == '__main__':
    main()
